#!/usr/bin/env python3
"""
Convert standards.xlsx to markdown table in docs/standards.md.

Usage:
    python script/excel_to_markdown.py
    # or from project root with venv:
    .venv/bin/python script/excel_to_markdown.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed.", file=sys.stderr)
    print("Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def wrap_text(text: str, max_length: int = 40) -> str:
    """
    Wrap text at word boundaries to max_length, inserting <br> tags.
    Only wraps if text is longer than max_length.
    """
    if len(text) <= max_length:
        return text
    
    words = text.split()
    lines = []
    current_line = []
    current_length = 0
    
    for word in words:
        word_length = len(word)
        # +1 for space before word (except first word)
        needed_length = word_length + (1 if current_line else 0)
        
        if current_length + needed_length <= max_length:
            current_line.append(word)
            current_length += needed_length
        else:
            # Start a new line
            if current_line:
                lines.append(" ".join(current_line))
            current_line = [word]
            current_length = word_length
    
    # Add the last line
    if current_line:
        lines.append(" ".join(current_line))
    
    return "<br>".join(lines)


def excel_to_markdown(excel_path: Path, sheet_name: str, output_path: Path):
    """Convert Excel sheet to markdown table."""
    
    # Load workbook and sheet
    wb = load_workbook(excel_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in {excel_path}", file=sys.stderr)
        print(f"Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # Extract all rows
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print(f"Error: Sheet '{sheet_name}' is empty", file=sys.stderr)
        sys.exit(1)
    
    # Build markdown content
    markdown_lines = [
        "# Relevant Standards",
        "",
        "> **Note:** This page is auto-generated from `doc/standards.xlsx`. ",
        "> To update, modify the Excel file and run: `.venv/bin/python script/excel_to_markdown.py`",
        "",
        "> **Note:** Tables on this page are designed to display at full browser width. On smaller screens, use horizontal scrolling to view all columns.",
        "",
    ]
    
    # Add section header based on sheet name
    markdown_lines.append(f"## {sheet_name}")
    markdown_lines.append("")
    
    # Get headers (first row)
    headers = rows[0]
    
    # Filter out None/empty headers
    num_cols = len([h for h in headers if h])
    
    # Build markdown table
    # Header row
    header_row = "| " + " | ".join(str(h or "") for h in headers[:num_cols]) + " |"
    markdown_lines.append(header_row)
    
    # Separator row (align left for all columns)
    separator = "|" + "|".join(":---" for _ in range(num_cols)) + "|"
    markdown_lines.append(separator)
    
    # Data rows
    for row in rows[1:]:
        # Skip completely empty rows
        if not any(row):
            continue
        
        # Convert cells to strings, handle None values
        cells = [str(cell if cell is not None else "") for cell in row[:num_cols]]
        
        # Process each cell: wrap long text (comments), escape special chars
        processed_cells = []
        for cell in cells:
            # Replace newlines with <br> first
            cell = cell.replace("\n", "<br>")
            # Wrap long text (likely comments/descriptions) at 40 chars
            if len(cell) > 80:  # Only wrap if longer than 80 chars
                cell = wrap_text(cell, max_length=40)
            # Escape pipe characters
            cell = cell.replace("|", "\\|")
            processed_cells.append(cell)
        
        row_text = "| " + " | ".join(processed_cells) + " |"
        markdown_lines.append(row_text)
    
    markdown_lines.append("")
    markdown_lines.append("> **Note:** For the complete table of standards and specifications, refer to the [relevant-standards.md](https://github.com/APTITUDE-Consortium/aptitude-eudi-wallet-rfcs/blob/main/doc/relevant-standards.md) file in the repository documentation folder.")
    markdown_lines.append("")
    
    # Write to output file
    output_path.write_text("\n".join(markdown_lines), encoding="utf-8")
    
    print(f"✓ Converted {len(rows)-1} rows from '{sheet_name}' to {output_path}")


def main():
    # Setup paths
    project_root = Path(__file__).parent.parent
    excel_path = project_root / "doc" / "standards.xlsx"
    output_path = project_root / "docs" / "standards.md"
    sheet_name = "Identification of standards&TS"
    
    # Validate input file exists
    if not excel_path.exists():
        print(f"Error: {excel_path} not found", file=sys.stderr)
        sys.exit(1)
    
    # Convert
    excel_to_markdown(excel_path, sheet_name, output_path)


if __name__ == "__main__":
    main()
